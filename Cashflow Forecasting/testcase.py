# test_accuracy.py
import pandas as pd
import numpy as np
import joblib
import os
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score, mean_absolute_percentage_error
import warnings
warnings.filterwarnings('ignore')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class AccuracyTester:
    """
    Comprehensive testing framework for evaluating AlcoBev forecasting model accuracy
    """

    def __init__(self, data_file='alcobev_europe_sales_data.csv', models_dir='models'):
        self.data_file = data_file
        self.models_dir = models_dir
        self.sales_model_path = os.path.join(models_dir, 'sales_forecasting_model.pkl')
        self.cogs_model_path = os.path.join(models_dir, 'cogs_forecasting_model.pkl')
        self.volume_model_path = os.path.join(models_dir, 'volume_forecasting_model.pkl')

        # Load data and models
        self.df = None
        self.sales_pipeline = None
        self.cogs_pipeline = None
        self.volume_pipeline = None
        self.test_results = {}

    def load_data_and_models(self):
        """Load the dataset and trained models"""
        print("Loading data and models...")

        # Load data
        try:
            self.df = pd.read_csv(self.data_file)
            self.df['Date'] = pd.to_datetime(self.df['Date'])
            print(f"✓ Data loaded: {len(self.df)} records")
        except FileNotFoundError:
            raise FileNotFoundError(f"Data file {self.data_file} not found. Run generate_data.py first.")

        # Load models
        try:
            if os.path.exists(self.sales_model_path):
                self.sales_pipeline = joblib.load(self.sales_model_path)
                print("✓ Sales model loaded successfully")
            else:
                raise FileNotFoundError(f"Sales model not found at {self.sales_model_path}")

            if os.path.exists(self.cogs_model_path):
                self.cogs_pipeline = joblib.load(self.cogs_model_path)
                print("✓ COGS model loaded successfully")
            else:
                raise FileNotFoundError(f"COGS model not found at {self.cogs_model_path}")

            if os.path.exists(self.volume_model_path):
                self.volume_pipeline = joblib.load(self.volume_model_path)
                print("✓ Volume model loaded successfully")
            else:
                raise FileNotFoundError(f"Volume model not found at {self.volume_model_path}")

        except Exception as e:
            raise Exception(f"Error loading models: {e}")

    def prepare_features(self, df):
        """Prepare features for Sales and COGS prediction (same as in training)"""
        df = df.copy()

        # Create time-based features
        df['day_of_week'] = df['Date'].dt.dayofweek
        df['month'] = df['Date'].dt.month
        df['year'] = df['Date'].dt.year
        df['day_of_year'] = df['Date'].dt.dayofyear
        df['week_of_year'] = df['Date'].dt.isocalendar().week.astype(int)

        # Create lagged features
        df = df.sort_values(by=['Country', 'Channel', 'Product_Category', 'Date'])
        df['Net_Sales_Revenue_EUR_lag1'] = df.groupby(['Country', 'Channel', 'Product_Category'])['Net_Sales_Revenue_EUR'].shift(1)
        df['COGS_EUR_lag1'] = df.groupby(['Country', 'Channel', 'Product_Category'])['COGS_EUR'].shift(1)

        # Drop rows with NaNs
        df.dropna(inplace=True)

        # Define feature order (must match training)
        feature_order = [
            'Net_Sales_Volume_Litres', 'Marketing_Spend_EUR', 'Promotional_Event',
            'Consumer_Confidence_Index', 'Inflation_Rate_EUR', 'Avg_Temp_C',
            'Holiday_Indicator', 'Competitor_Activity_Index',
            'day_of_week', 'month', 'year', 'day_of_year', 'week_of_year',
            'Net_Sales_Revenue_EUR_lag1', 'COGS_EUR_lag1',
            'Country', 'Channel', 'Product_Category'
        ]

        return df[feature_order + ['Net_Sales_Revenue_EUR', 'COGS_EUR', 'Date']]

    def volume_features(self, df):
        """Prepare features for volume prediction (same as in training)"""
        df = df.copy()

        # Create time-based features
        df['day_of_week'] = df['Date'].dt.dayofweek
        df['month'] = df['Date'].dt.month
        df['year'] = df['Date'].dt.year
        df['day_of_year'] = df['Date'].dt.dayofyear
        df['week_of_year'] = df['Date'].dt.isocalendar().week.astype(int)

        # Create lagged features
        df = df.sort_values(by=['Country', 'Channel', 'Product_Category', 'Date'])
        df['Net_Sales_Volume_Litres_lag1'] = df.groupby(['Country', 'Channel', 'Product_Category'])['Net_Sales_Volume_Litres'].shift(1)

        # Drop rows with NaNs
        df.dropna(inplace=True)

        # Define feature order (must match training)
        feature_order = [
            'Marketing_Spend_EUR', 'Promotional_Event',
            'Consumer_Confidence_Index', 'Inflation_Rate_EUR', 'Avg_Temp_C',
            'Holiday_Indicator', 'Competitor_Activity_Index',
            'day_of_week', 'month', 'year', 'day_of_year', 'week_of_year',
            'Net_Sales_Volume_Litres_lag1',
            'Country', 'Channel', 'Product_Category'
        ]

        return df[feature_order + ['Net_Sales_Volume_Litres', 'Date']]

    def calculate_metrics(self, y_true, y_pred, model_name):
        """Calculate comprehensive accuracy metrics"""
        metrics = {
            'MAE': mean_absolute_error(y_true, y_pred),
            'RMSE': np.sqrt(mean_squared_error(y_true, y_pred)),
            'R2': r2_score(y_true, y_pred),
            'MAPE': mean_absolute_percentage_error(y_true, y_pred) * 100,
            'Mean_Actual': np.mean(y_true),
            'Mean_Predicted': np.mean(y_pred),
            'Std_Actual': np.std(y_true),
            'Std_Predicted': np.std(y_pred)
        }

        # Additional custom metrics
        metrics['Bias'] = np.mean(y_pred - y_true)  # Average prediction bias
        metrics['Bias_Percentage'] = (metrics['Bias'] / metrics['Mean_Actual']) * 100
        metrics['Accuracy_90'] = np.mean(np.abs((y_true - y_pred) / y_true) <= 0.1) * 100  # % within 10%
        metrics['Accuracy_80'] = np.mean(np.abs((y_true - y_pred) / y_true) <= 0.2) * 100  # % within 20%

        return metrics

    def test_holdout_accuracy(self, test_split_date='2024-01-01'):
        """Test accuracy on holdout test set"""
        print(f"\n=== HOLDOUT TEST ACCURACY (Test data from {test_split_date}) ===")

        # Prepare data
        df_prepared = self.prepare_features(self.df)
        df_v_prepared = self.volume_features(self.df)
        test_split_date = pd.to_datetime(test_split_date)

        # Split data
        test_df = df_prepared[df_prepared['Date'] >= test_split_date].copy()
        test_v_df = df_v_prepared[df_v_prepared['Date'] >= test_split_date].copy()

        if len(test_df) == 0 or len(test_v_df) == 0:
            print("❌ No test data available for the specified date")
            return

        print(f"Test set size: {len(test_df)} records (Sales/COGS), {len(test_v_df)} records (Volume)")

        # Prepare features and targets
        feature_cols = [col for col in test_df.columns if col not in ['Net_Sales_Revenue_EUR', 'COGS_EUR', 'Date']]
        volume_feature_cols = [col for col in test_v_df.columns if col not in ['Net_Sales_Volume_Litres', 'Date']]

        X_test = test_df[feature_cols]
        vX_test = test_v_df[volume_feature_cols]

        y_test_sales = test_df['Net_Sales_Revenue_EUR']
        y_test_cogs = test_df['COGS_EUR']
        y_test_volume = test_v_df['Net_Sales_Volume_Litres']

        # Make predictions
        y_pred_sales = self.sales_pipeline.predict(X_test)
        y_pred_cogs = self.cogs_pipeline.predict(X_test)
        y_pred_volume = self.volume_pipeline.predict(vX_test)

        # Calculate metrics
        sales_metrics = self.calculate_metrics(y_test_sales, y_pred_sales, 'Sales')
        cogs_metrics = self.calculate_metrics(y_test_cogs, y_pred_cogs, 'COGS')
        volume_metrics = self.calculate_metrics(y_test_volume, y_pred_volume, 'Volume')

        # Store results
        self.test_results['holdout'] = {
            'sales': sales_metrics,
            'cogs': cogs_metrics,
            'volume': volume_metrics,
            'test_data': test_df,
            'test_volume_data': test_v_df,  # Store volume test data separately
            'predictions': {
                'sales': y_pred_sales,
                'cogs': y_pred_cogs,
                'volume': y_pred_volume
            }
        }

        # Print results
        self.print_metrics_table(sales_metrics, cogs_metrics, volume_metrics, "HOLDOUT TEST")

        return sales_metrics, cogs_metrics, volume_metrics

    def test_rolling_window_accuracy(self, window_days=30, n_windows=5):
        """Test accuracy using rolling windows"""
        print(f"\n=== ROLLING WINDOW ACCURACY ({n_windows} windows of {window_days} days) ===")

        df_prepared = self.prepare_features(self.df)
        df_prepared = df_prepared.sort_values('Date')
        df_volume = self.volume_features(self.df)
        df_volume = df_volume.sort_values('Date')

        # Get the last n_windows worth of data
        end_date = df_prepared['Date'].max()
        start_date = end_date - timedelta(days=window_days * n_windows)

        recent_data = df_prepared[df_prepared['Date'] >= start_date].copy()
        recent_volume_data = df_volume[df_volume['Date'] >= start_date].copy()

        rolling_results = {'sales': [], 'cogs': [], 'volume': []}

        for i in range(n_windows):
            window_start = start_date + timedelta(days=i * window_days)
            window_end = window_start + timedelta(days=window_days)

            window_data = recent_data[
                (recent_data['Date'] >= window_start) & 
                (recent_data['Date'] < window_end)
            ].copy()

            window_volume_data = recent_volume_data[
                (recent_volume_data['Date'] >= window_start) & 
                (recent_volume_data['Date'] < window_end)
            ].copy()

            if len(window_data) < 10 or len(window_volume_data) < 10:  # Skip if too few samples
                continue

            # Prepare features
            feature_cols = [col for col in window_data.columns if col not in ['Net_Sales_Revenue_EUR', 'COGS_EUR', 'Date']]
            volume_feature_cols = [col for col in window_volume_data.columns if col not in ['Net_Sales_Volume_Litres', 'Date']]

            X_window = window_data[feature_cols]
            X_window_volume = window_volume_data[volume_feature_cols]

            y_sales = window_data['Net_Sales_Revenue_EUR']
            y_cogs = window_data['COGS_EUR']
            y_volume = window_volume_data['Net_Sales_Volume_Litres']

            # Make predictions
            pred_sales = self.sales_pipeline.predict(X_window)
            pred_cogs = self.cogs_pipeline.predict(X_window)
            pred_volume = self.volume_pipeline.predict(X_window_volume)

            # Calculate metrics
            sales_metrics = self.calculate_metrics(y_sales, pred_sales, f'Sales_Window_{i+1}')
            cogs_metrics = self.calculate_metrics(y_cogs, pred_cogs, f'COGS_Window_{i+1}')
            volume_metrics = self.calculate_metrics(y_volume, pred_volume, f'Volume_Window_{i+1}')

            rolling_results['sales'].append(sales_metrics)
            rolling_results['cogs'].append(cogs_metrics)
            rolling_results['volume'].append(volume_metrics)

            print(f"Window {i+1} ({window_start.strftime('%Y-%m-%d')} to {window_end.strftime('%Y-%m-%d')}): "
                  f"Sales MAPE: {sales_metrics['MAPE']:.2f}%, COGS MAPE: {cogs_metrics['MAPE']:.2f}%, Volume MAPE: {volume_metrics['MAPE']:.2f}%")

        # Calculate average metrics across windows
        if rolling_results['sales']:  # Check if we have results
            avg_sales_metrics = {metric: np.mean([w[metric] for w in rolling_results['sales']]) 
                               for metric in rolling_results['sales'][0].keys()}
            avg_cogs_metrics = {metric: np.mean([w[metric] for w in rolling_results['cogs']]) 
                              for metric in rolling_results['cogs'][0].keys()}
            avg_volume_metrics = {metric: np.mean([w[metric] for w in rolling_results['volume']]) 
                                for metric in rolling_results['volume'][0].keys()}

            self.test_results['rolling'] = {
                'sales': avg_sales_metrics,
                'cogs': avg_cogs_metrics,
                'volume': avg_volume_metrics,
                'individual_windows': rolling_results
            }

            print(f"\n📊 ROLLING WINDOW AVERAGES:")
            self.print_metrics_table(avg_sales_metrics, avg_cogs_metrics, avg_volume_metrics, "ROLLING AVERAGE")

            return avg_sales_metrics, avg_cogs_metrics, avg_volume_metrics
        else:
            print("❌ No rolling window results available")
            return None, None, None

    def test_segment_accuracy(self):
        """Test accuracy by different segments (Country, Channel, Product)"""
        print(f"\n=== SEGMENT-WISE ACCURACY ===")

        df_prepared = self.prepare_features(self.df)
        df_volume = self.volume_features(self.df)

        # Use last 6 months for segment testing
        recent_date = df_prepared['Date'].max() - timedelta(days=180)
        recent_volume_date = df_volume['Date'].max() - timedelta(days=180)

        test_data = df_prepared[df_prepared['Date'] >= recent_date].copy()
        test_volume_data = df_volume[df_volume['Date'] >= recent_volume_date].copy()

        segments = {
            'Country': test_data['Country'].unique(),
            'Channel': test_data['Channel'].unique(),  
            'Product_Category': test_data['Product_Category'].unique()
        }

        segment_results = {}

        for segment_type, segment_values in segments.items():
            print(f"\n📈 {segment_type} Segment Analysis:")
            segment_results[segment_type] = {}

            for segment_value in segment_values:
                segment_data = test_data[test_data[segment_type] == segment_value].copy()
                segment_volume_data = test_volume_data[test_volume_data[segment_type] == segment_value].copy()

                if len(segment_data) < 20 or len(segment_volume_data) < 20:  # Skip if too few samples
                    continue

                # Prepare features
                feature_cols = [col for col in segment_data.columns if col not in ['Net_Sales_Revenue_EUR', 'COGS_EUR', 'Date']]
                volume_feature_cols = [col for col in segment_volume_data.columns if col not in ['Net_Sales_Volume_Litres', 'Date']]

                X_segment = segment_data[feature_cols]
                X_segment_volume = segment_volume_data[volume_feature_cols]

                y_sales = segment_data['Net_Sales_Revenue_EUR']
                y_cogs = segment_data['COGS_EUR']
                y_volume = segment_volume_data['Net_Sales_Volume_Litres']

                # Make predictions
                pred_sales = self.sales_pipeline.predict(X_segment)
                pred_cogs = self.cogs_pipeline.predict(X_segment)
                pred_volume = self.volume_pipeline.predict(X_segment_volume)

                # Calculate metrics
                sales_metrics = self.calculate_metrics(y_sales, pred_sales, f'Sales_{segment_value}')
                cogs_metrics = self.calculate_metrics(y_cogs, pred_cogs, f'COGS_{segment_value}')
                volume_metrics = self.calculate_metrics(y_volume, pred_volume, f'Volume_{segment_value}')

                segment_results[segment_type][segment_value] = {
                    'sales': sales_metrics,
                    'cogs': cogs_metrics,
                    'volume': volume_metrics
                }

                print(f"  {segment_value:15} | Sales MAPE: {sales_metrics['MAPE']:6.2f}% | COGS MAPE: {cogs_metrics['MAPE']:6.2f}% | Volume MAPE: {volume_metrics['MAPE']:6.2f}% | Sales R²: {sales_metrics['R2']:.3f} | Volume R²: {volume_metrics['R2']:.3f}")

        self.test_results['segments'] = segment_results
        return segment_results

    def print_metrics_table(self, sales_metrics, cogs_metrics, volume_metrics, title):
        """Print formatted metrics table"""
        print(f"\n📊 {title} RESULTS:")
        print("=" * 100)
        print(f"{'Metric':<20} | {'Sales Model':<15} | {'COGS Model':<15} | {'Volume Model':<15}")
        print("-" * 100)
        print(f"{'MAE':<20} | {sales_metrics.get('MAE', float('nan')):>13,.0f} | {cogs_metrics.get('MAE', float('nan')):>13,.0f} | {volume_metrics.get('MAE', float('nan')):>13,.0f}")
        print(f"{'RMSE':<20} | {sales_metrics.get('RMSE', float('nan')):>13,.0f} | {cogs_metrics.get('RMSE', float('nan')):>13,.0f} | {volume_metrics.get('RMSE', float('nan')):>13,.0f}")
        print(f"{'MAPE (%)':<20} | {sales_metrics.get('MAPE', float('nan')):>13.2f} | {cogs_metrics.get('MAPE', float('nan')):>13.2f} | {volume_metrics.get('MAPE', float('nan')):>13.2f}")
        print(f"{'R² Score':<20} | {sales_metrics.get('R2', float('nan')):>13.3f} | {cogs_metrics.get('R2', float('nan')):>13.3f} | {volume_metrics.get('R2', float('nan')):>13.3f}")
        print(f"{'Bias':<20} | {sales_metrics.get('Bias', float('nan')):>13,.0f} | {cogs_metrics.get('Bias', float('nan')):>13,.0f} | {volume_metrics.get('Bias', float('nan')):>13,.0f}")
        print(f"{'Bias (%)':<20} | {sales_metrics.get('Bias_Percentage', float('nan')):>13.2f} | {cogs_metrics.get('Bias_Percentage', float('nan')):>13.2f} | {volume_metrics.get('Bias_Percentage', float('nan')):>13.2f}")
        print(f"{'Within 10% (%)':<20} | {sales_metrics.get('Accuracy_90', float('nan')):>13.1f} | {cogs_metrics.get('Accuracy_90', float('nan')):>13.1f} | {volume_metrics.get('Accuracy_90', float('nan')):>13.1f}")
        print(f"{'Within 20% (%)':<20} | {sales_metrics.get('Accuracy_80', float('nan')):>13.1f} | {cogs_metrics.get('Accuracy_80', float('nan')):>13.1f} | {volume_metrics.get('Accuracy_80', float('nan')):>13.1f}")
        print("=" * 100)

    def generate_accuracy_report(self):
        """Generate comprehensive accuracy report"""
        print("\n" + "="*100)
        print(" " * 30 + "ALCOBEV MODEL ACCURACY REPORT")
        print("="*100)

        # Overall assessment
        if 'holdout' in self.test_results:
            sales_mape = self.test_results['holdout']['sales']['MAPE']
            cogs_mape = self.test_results['holdout']['cogs']['MAPE']
            volume_mape = self.test_results['holdout']['volume']['MAPE']
            sales_r2 = self.test_results['holdout']['sales']['R2']
            cogs_r2 = self.test_results['holdout']['cogs']['R2']
            volume_r2 = self.test_results['holdout']['volume']['R2']

            print(f"\n🎯 OVERALL ASSESSMENT:")
            print(f"  Sales Model: MAPE = {sales_mape:.2f}%, R² = {sales_r2:.3f}")
            print(f"  COGS Model: MAPE = {cogs_mape:.2f}%, R² = {cogs_r2:.3f}")
            print(f"  Volume Model: MAPE = {volume_mape:.2f}%, R² = {volume_r2:.3f}")

            # Accuracy classification
            def classify_accuracy(mape, r2):
                if mape <= 5 and r2 >= 0.9:
                    return "EXCELLENT 🟢"
                elif mape <= 10 and r2 >= 0.8:
                    return "GOOD 🟡"
                elif mape <= 20 and r2 >= 0.6:
                    return "FAIR 🟠"
                else:
                    return "NEEDS IMPROVEMENT 🔴"

            sales_rating = classify_accuracy(sales_mape, sales_r2)
            cogs_rating = classify_accuracy(cogs_mape, cogs_r2)
            volume_rating = classify_accuracy(volume_mape, volume_r2)

            print(f"\nMODEL RATINGS:")
            print(f"  Sales Model: {sales_rating}")
            print(f"  COGS Model: {cogs_rating}")
            print(f"  Volume Model: {volume_rating}")

            # Recommendations
            print(f"\nRECOMMENDATIONS:")
            if sales_mape > 15:
                print("  Sales model may benefit from additional features or hyperparameter tuning")
            if cogs_mape > 15:
                print("  COGS model may need more training data or feature engineering")
            if volume_mape > 15:
                print("  Volume model may benefit from additional features or hyperparameter tuning")
            if sales_r2 < 0.7 or cogs_r2 < 0.7 or volume_r2 < 0.7:
                print("  Consider ensemble methods or more sophisticated algorithms")

            print("  Regular model retraining recommended (monthly/quarterly)")
            print("  Monitor prediction accuracy on new data continuously")
            print("  Consider A/B testing predictions against business forecasts")

        return self.test_results

    def create_visualizations(self):
        """Create accuracy visualization plots"""
        if 'holdout' not in self.test_results:
            print("No holdout test results available for visualization")
            return

        # Get test data and predictions
        test_data = self.test_results['holdout']['test_data']
        test_volume_data = self.test_results['holdout']['test_volume_data']
        sales_pred = self.test_results['holdout']['predictions']['sales']
        cogs_pred = self.test_results['holdout']['predictions']['cogs']
        volume_pred = self.test_results['holdout']['predictions']['volume']

        # SALES VISUALIZATION
        fig, axes = plt.subplots(1, 2, figsize=(15, 5))
        fig.suptitle('Sales Model Accuracy Visualization', fontsize=16, fontweight='bold')

        # Sales: Actual vs Predicted scatter
        axes[0].scatter(test_data['Net_Sales_Revenue_EUR'], sales_pred, alpha=0.6, color='green')
        axes[0].plot([test_data['Net_Sales_Revenue_EUR'].min(), test_data['Net_Sales_Revenue_EUR'].max()],
                    [test_data['Net_Sales_Revenue_EUR'].min(), test_data['Net_Sales_Revenue_EUR'].max()],
                    'r--', lw=2)
        axes[0].set_xlabel('Actual Sales Revenue (€)')
        axes[0].set_ylabel('Predicted Sales Revenue (€)')
        axes[0].set_title('Sales: Actual vs Predicted')

        # Sales residuals
        sales_residuals = test_data['Net_Sales_Revenue_EUR'] - sales_pred
        axes[1].scatter(sales_pred, sales_residuals, alpha=0.6, color='green')
        axes[1].axhline(y=0, color='r', linestyle='--')
        axes[1].set_xlabel('Predicted Sales Revenue (€)')
        axes[1].set_ylabel('Residuals (€)')
        axes[1].set_title('Sales Residuals Plot')

        plt.tight_layout()
        plt.savefig('sales_accuracy_plots.png', dpi=300, bbox_inches='tight')
        plt.show()

        # COGS VISUALIZATION
        fig, axes = plt.subplots(1, 2, figsize=(15, 5))
        fig.suptitle('COGS Model Accuracy Visualization', fontsize=16, fontweight='bold')

        # COGS: Actual vs Predicted scatter
        axes[0].scatter(test_data['COGS_EUR'], cogs_pred, alpha=0.6, color='red')
        axes[0].plot([test_data['COGS_EUR'].min(), test_data['COGS_EUR'].max()],
                    [test_data['COGS_EUR'].min(), test_data['COGS_EUR'].max()],
                    'r--', lw=2)
        axes[0].set_xlabel('Actual COGS (€)')
        axes[0].set_ylabel('Predicted COGS (€)')
        axes[0].set_title('COGS: Actual vs Predicted')

        # COGS residuals
        cogs_residuals = test_data['COGS_EUR'] - cogs_pred
        axes[1].scatter(cogs_pred, cogs_residuals, alpha=0.6, color='red')
        axes[1].axhline(y=0, color='r', linestyle='--')
        axes[1].set_xlabel('Predicted COGS (€)')
        axes[1].set_ylabel('Residuals (€)')
        axes[1].set_title('COGS Residuals Plot')

        plt.tight_layout()
        plt.savefig('cogs_accuracy_plots.png', dpi=300, bbox_inches='tight')
        plt.show()

        # VOLUME VISUALIZATION - Fixed to use correct test data
        fig, axes = plt.subplots(1, 2, figsize=(15, 5))
        fig.suptitle('Volume Model Accuracy Visualization', fontsize=16, fontweight='bold')

        # Volume: Actual vs Predicted scatter - Using test_volume_data instead of test_data
        axes[0].scatter(test_volume_data['Net_Sales_Volume_Litres'], volume_pred, alpha=0.6, color='blue')
        axes[0].plot([test_volume_data['Net_Sales_Volume_Litres'].min(), test_volume_data['Net_Sales_Volume_Litres'].max()],
                    [test_volume_data['Net_Sales_Volume_Litres'].min(), test_volume_data['Net_Sales_Volume_Litres'].max()],
                    'r--', lw=2)
        axes[0].set_xlabel('Actual Volume (Litres)')
        axes[0].set_ylabel('Predicted Volume (Litres)')
        axes[0].set_title('Volume: Actual vs Predicted')

        # Volume residuals
        volume_residuals = test_volume_data['Net_Sales_Volume_Litres'] - volume_pred
        axes[1].scatter(volume_pred, volume_residuals, alpha=0.6, color='blue')
        axes[1].axhline(y=0, color='r', linestyle='--')
        axes[1].set_xlabel('Predicted Volume (Litres)')
        axes[1].set_ylabel('Residuals (Litres)')
        axes[1].set_title('Volume Residuals Plot')

        plt.tight_layout()
        plt.savefig('volume_accuracy_plots.png', dpi=300, bbox_inches='tight')
        plt.show()

    def save_results_to_excel(self, filename=None):
        """Save all test results to a comprehensive Excel file"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"AlcoBev_Model_Accuracy_Report_{timestamp}.xlsx"

        print(f"\n💾 Saving results to Excel file: {filename}")

        # Create Excel writer object
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 1. Executive Summary Sheet
            self._create_executive_summary_sheet(writer)

            # 2. Holdout Test Results
            if 'holdout' in self.test_results:
                self._create_holdout_test_sheet(writer)

            # 3. Rolling Window Results
            if 'rolling' in self.test_results:
                self._create_rolling_window_sheet(writer)

            # 4. Segment Analysis Results
            if 'segments' in self.test_results:
                self._create_segment_analysis_sheet(writer)

            # 5. Detailed Predictions (if available)
            if 'holdout' in self.test_results:
                self._create_detailed_predictions_sheet(writer)

            # 6. Model Information Sheet
            self._create_model_info_sheet(writer)

        # Apply formatting to the Excel file
        self._format_excel_file(filename)

        print(f"✅ Excel report saved successfully: {filename}")
        return filename

    def _create_executive_summary_sheet(self, writer):
        """Create executive summary sheet"""
        summary_data = []

        # Test run information
        summary_data.append(['Test Run Information', '', '', ''])
        summary_data.append(['Report Generated', datetime.now().strftime("%Y-%m-%d %H:%M:%S"), '', ''])
        summary_data.append(['Data File', self.data_file, '', ''])
        summary_data.append(['Models Directory', self.models_dir, '', ''])
        summary_data.append(['', '', '', ''])

        # Overall accuracy metrics
        if 'holdout' in self.test_results:
            summary_data.append(['Overall Model Performance (Holdout Test)', '', '', ''])
            summary_data.append(['Metric', 'Sales Model', 'COGS Model', 'Volume Model'])

            sales_metrics = self.test_results['holdout']['sales']
            cogs_metrics = self.test_results['holdout']['cogs']
            volume_metrics = self.test_results['holdout']['volume']

            summary_data.append(['MAE', f"{sales_metrics['MAE']:,.0f}", f"{cogs_metrics['MAE']:,.0f}", f"{volume_metrics['MAE']:,.0f}"])
            summary_data.append(['RMSE', f"{sales_metrics['RMSE']:,.0f}", f"{cogs_metrics['RMSE']:,.0f}", f"{volume_metrics['RMSE']:,.0f}"])
            summary_data.append(['MAPE (%)', f"{sales_metrics['MAPE']:.2f}", f"{cogs_metrics['MAPE']:.2f}", f"{volume_metrics['MAPE']:.2f}"])
            summary_data.append(['R² Score', f"{sales_metrics['R2']:.3f}", f"{cogs_metrics['R2']:.3f}", f"{volume_metrics['R2']:.3f}"])
            summary_data.append(['Bias (%)', f"{sales_metrics['Bias_Percentage']:.2f}", f"{cogs_metrics['Bias_Percentage']:.2f}", f"{volume_metrics['Bias_Percentage']:.2f}"])
            summary_data.append(['Within 10% Accuracy (%)', f"{sales_metrics['Accuracy_90']:.1f}", f"{cogs_metrics['Accuracy_90']:.1f}", f"{volume_metrics['Accuracy_90']:.1f}"])
            summary_data.append(['Within 20% Accuracy (%)', f"{sales_metrics['Accuracy_80']:.1f}", f"{cogs_metrics['Accuracy_80']:.1f}", f"{volume_metrics['Accuracy_80']:.1f}"])
            summary_data.append(['', '', '', ''])

            # Model ratings
            def classify_accuracy(mape, r2):
                if mape <= 5 and r2 >= 0.9:
                    return "EXCELLENT"
                elif mape <= 10 and r2 >= 0.8:
                    return "GOOD"
                elif mape <= 20 and r2 >= 0.6:
                    return "FAIR"
                else:
                    return "NEEDS IMPROVEMENT"

            sales_rating = classify_accuracy(sales_metrics['MAPE'], sales_metrics['R2'])
            cogs_rating = classify_accuracy(cogs_metrics['MAPE'], cogs_metrics['R2'])
            volume_rating = classify_accuracy(volume_metrics['MAPE'], volume_metrics['R2'])

            summary_data.append(['Model Assessment', '', '', ''])
            summary_data.append(['Sales Model Rating', sales_rating, '', ''])
            summary_data.append(['COGS Model Rating', cogs_rating, '', ''])
            summary_data.append(['Volume Model Rating', volume_rating, '', ''])
            summary_data.append(['', '', '', ''])

            # Rolling window performance (if available)
            if 'rolling' in self.test_results:
                summary_data.append(['Rolling Window Performance (Average)', '', '', ''])
                rolling_sales = self.test_results['rolling']['sales']
                rolling_cogs = self.test_results['rolling']['cogs']
                rolling_volume = self.test_results['rolling']['volume']

                summary_data.append(['Sales MAPE (%)', f"{rolling_sales['MAPE']:.2f}", '', ''])
                summary_data.append(['COGS MAPE (%)', f"{rolling_cogs['MAPE']:.2f}", '', ''])
                summary_data.append(['Volume MAPE (%)', f"{rolling_volume['MAPE']:.2f}", '', ''])
                summary_data.append(['Sales R²', f"{rolling_sales['R2']:.3f}", '', ''])
                summary_data.append(['COGS R²', f"{rolling_cogs['R2']:.3f}", '', ''])
                summary_data.append(['Volume R²', f"{rolling_volume['R2']:.3f}", '', ''])
                summary_data.append(['', '', '', ''])

            # Recommendations
            summary_data.append(['Recommendations', '', '', ''])
            recommendations = self._generate_recommendations()
            for i, rec in enumerate(recommendations, 1):
                summary_data.append([f'{i}.', rec, '', ''])

        # Convert to DataFrame and save
        df_summary = pd.DataFrame(summary_data, columns=['Category', 'Value', 'Additional', 'Extra'])
        df_summary.to_excel(writer, sheet_name='Executive Summary', index=False)

    def _create_holdout_test_sheet(self, writer):
        """Create holdout test results sheet"""
        holdout_results = self.test_results['holdout']

        # Metrics comparison
        sales_metrics = holdout_results['sales']
        cogs_metrics = holdout_results['cogs']
        volume_metrics = holdout_results['volume']

        metrics_data = []
        for metric in sales_metrics.keys():
            if metric in ['Mean_Actual', 'Mean_Predicted', 'Std_Actual', 'Std_Predicted']:
                sales_val = f"{sales_metrics[metric]:,.0f}"
                cogs_val = f"{cogs_metrics[metric]:,.0f}"
                volume_val = f"{volume_metrics[metric]:,.0f}"
            elif metric in ['MAPE', 'Bias_Percentage', 'Accuracy_90', 'Accuracy_80']:
                sales_val = f"{sales_metrics[metric]:.2f}%"
                cogs_val = f"{cogs_metrics[metric]:.2f}%"
                volume_val = f"{volume_metrics[metric]:.2f}%"
            elif metric in ['R2']:
                sales_val = f"{sales_metrics[metric]:.4f}"
                cogs_val = f"{cogs_metrics[metric]:.4f}"
                volume_val = f"{volume_metrics[metric]:.4f}"
            else:
                sales_val = f"{sales_metrics[metric]:,.0f}"
                cogs_val = f"{cogs_metrics[metric]:,.0f}"
                volume_val = f"{volume_metrics[metric]:,.0f}"

            metrics_data.append([metric, sales_val, cogs_val, volume_val])

        df_metrics = pd.DataFrame(metrics_data, columns=['Metric', 'Sales Model', 'COGS Model', 'Volume Model'])
        df_metrics.to_excel(writer, sheet_name='Holdout Test Results', index=False)

    def _create_rolling_window_sheet(self, writer):
        """Create rolling window analysis sheet"""
        if 'rolling' not in self.test_results:
            return

        rolling_results = self.test_results['rolling']

        # Individual window results
        windows_data = []
        individual_windows = rolling_results['individual_windows']

        if individual_windows['sales']:  # Check if we have results
            for i, (sales_window, cogs_window, volume_window) in enumerate(zip(
                individual_windows['sales'], 
                individual_windows['cogs'], 
                individual_windows['volume']), 1):

                windows_data.append([
                    f'Window {i}',
                    f"{sales_window['MAPE']:.2f}%",
                    f"{sales_window['R2']:.3f}",
                    f"{sales_window['MAE']:,.0f}",
                    f"{cogs_window['MAPE']:.2f}%", 
                    f"{cogs_window['R2']:.3f}",
                    f"{cogs_window['MAE']:,.0f}",
                    f"{volume_window['MAPE']:.2f}%",
                    f"{volume_window['R2']:.3f}",
                    f"{volume_window['MAE']:,.0f}"
                ])

            # Add averages
            avg_sales = rolling_results['sales']
            avg_cogs = rolling_results['cogs'] 
            avg_volume = rolling_results['volume']

            windows_data.append([
                'AVERAGE',
                f"{avg_sales['MAPE']:.2f}%",
                f"{avg_sales['R2']:.3f}",
                f"{avg_sales['MAE']:,.0f}",
                f"{avg_cogs['MAPE']:.2f}%",
                f"{avg_cogs['R2']:.3f}",
                f"{avg_cogs['MAE']:,.0f}",
                f"{avg_volume['MAPE']:.2f}%",
                f"{avg_volume['R2']:.3f}",
                f"{avg_volume['MAE']:,.0f}"
            ])

        df_windows = pd.DataFrame(windows_data, columns=[
            'Window', 'Sales MAPE', 'Sales R²', 'Sales MAE',
            'COGS MAPE', 'COGS R²', 'COGS MAE', 
            'Volume MAPE', 'Volume R²', 'Volume MAE'
        ])
        df_windows.to_excel(writer, sheet_name='Rolling Window Analysis', index=False)

    def _create_segment_analysis_sheet(self, writer):
        """Create segment analysis sheet"""
        if 'segments' not in self.test_results:
            return

        segments_results = self.test_results['segments']

        segment_data = []
        for segment_type, segment_values in segments_results.items():
            segment_data.append([f'{segment_type} Analysis', '', '', '', '', '', '', '', '', ''])
            segment_data.append(['Segment', 'Sales MAPE (%)', 'Sales R²', 'Sales MAE', 
                               'COGS MAPE (%)', 'COGS R²', 'COGS MAE', 
                               'Volume MAPE (%)', 'Volume R²', 'Volume MAE'])

            for segment_value, metrics in segment_values.items():
                sales_metrics = metrics['sales']
                cogs_metrics = metrics['cogs']
                volume_metrics = metrics['volume']

                segment_data.append([
                    segment_value,
                    f"{sales_metrics['MAPE']:.2f}",
                    f"{sales_metrics['R2']:.3f}",
                    f"{sales_metrics['MAE']:,.0f}",
                    f"{cogs_metrics['MAPE']:.2f}",
                    f"{cogs_metrics['R2']:.3f}",
                    f"{cogs_metrics['MAE']:,.0f}",
                    f"{volume_metrics['MAPE']:.2f}",
                    f"{volume_metrics['R2']:.3f}",
                    f"{volume_metrics['MAE']:,.0f}"
                ])

            segment_data.append(['', '', '', '', '', '', '', '', '', ''])  # Empty row

        df_segments = pd.DataFrame(segment_data, columns=[
            'Segment', 'Sales MAPE (%)', 'Sales R²', 'Sales MAE',
            'COGS MAPE (%)', 'COGS R²', 'COGS MAE',
            'Volume MAPE (%)', 'Volume R²', 'Volume MAE'
        ])
        df_segments.to_excel(writer, sheet_name='Segment Analysis', index=False)

    def _create_detailed_predictions_sheet(self, writer):
        """Create detailed predictions sheet with actual vs predicted values"""
        if 'holdout' not in self.test_results:
            return

        test_data = self.test_results['holdout']['test_data']
        test_volume_data = self.test_results['holdout']['test_volume_data']
        sales_pred = self.test_results['holdout']['predictions']['sales']
        cogs_pred = self.test_results['holdout']['predictions']['cogs']
        volume_pred = self.test_results['holdout']['predictions']['volume']

        # Create detailed predictions DataFrame for Sales/COGS
        detailed_data = test_data[['Date', 'Country', 'Channel', 'Product_Category',
                                 'Net_Sales_Revenue_EUR', 'COGS_EUR']].copy()
        detailed_data['Predicted_Sales_EUR'] = sales_pred
        detailed_data['Predicted_COGS_EUR'] = cogs_pred
        detailed_data['Sales_Error_EUR'] = detailed_data['Net_Sales_Revenue_EUR'] - detailed_data['Predicted_Sales_EUR']
        detailed_data['COGS_Error_EUR'] = detailed_data['COGS_EUR'] - detailed_data['Predicted_COGS_EUR']
        detailed_data['Sales_Error_Pct'] = (detailed_data['Sales_Error_EUR'] / detailed_data['Net_Sales_Revenue_EUR']) * 100
        detailed_data['COGS_Error_Pct'] = (detailed_data['COGS_Error_EUR'] / detailed_data['COGS_EUR']) * 100

        # Create detailed predictions DataFrame for Volume
        volume_detailed = test_volume_data[['Date', 'Country', 'Channel', 'Product_Category',
                                          'Net_Sales_Volume_Litres']].copy()
        volume_detailed['Predicted_Volume'] = volume_pred
        volume_detailed['Volume_Error'] = volume_detailed['Net_Sales_Volume_Litres'] - volume_detailed['Predicted_Volume']
        volume_detailed['Volume_Error_Pct'] = (volume_detailed['Volume_Error'] / volume_detailed['Net_Sales_Volume_Litres']) * 100

        # Round numerical columns
        numerical_cols = ['Net_Sales_Revenue_EUR', 'COGS_EUR', 'Predicted_Sales_EUR', 
                         'Predicted_COGS_EUR', 'Sales_Error_EUR', 'COGS_Error_EUR']
        for col in numerical_cols:
            detailed_data[col] = detailed_data[col].round(2)
        detailed_data['Sales_Error_Pct'] = detailed_data['Sales_Error_Pct'].round(2)
        detailed_data['COGS_Error_Pct'] = detailed_data['COGS_Error_Pct'].round(2)

        # Round volume columns
        volume_numerical_cols = ['Net_Sales_Volume_Litres', 'Predicted_Volume', 'Volume_Error']
        for col in volume_numerical_cols:
            volume_detailed[col] = volume_detailed[col].round(2)
        volume_detailed['Volume_Error_Pct'] = volume_detailed['Volume_Error_Pct'].round(2)

        # Save to Excel
        detailed_data.to_excel(writer, sheet_name='Detailed Predictions - Sales COGS', index=False)
        volume_detailed.to_excel(writer, sheet_name='Detailed Predictions - Volume', index=False)

    def _create_model_info_sheet(self, writer):
        """Create model information sheet"""
        model_info = []

        model_info.append(['Model Information', ''])
        model_info.append(['', ''])
        model_info.append(['Sales Model Path', self.sales_model_path])
        model_info.append(['COGS Model Path', self.cogs_model_path])
        model_info.append(['Volume Model Path', self.volume_model_path])
        model_info.append(['', ''])

        # Model details (if available)
        if self.sales_pipeline:
            model_info.append(['Sales Model Details', ''])
            try:
                model_info.append(['Algorithm', str(type(self.sales_pipeline.named_steps['regressor']).__name__)])
                model_info.append(['Features Expected', str(self.sales_pipeline.named_steps['preprocessor'].n_features_in_)])
            except:
                model_info.append(['Algorithm', 'Information not available'])

        if self.cogs_pipeline:
            model_info.append(['', ''])
            model_info.append(['COGS Model Details', ''])
            try:
                model_info.append(['Algorithm', str(type(self.cogs_pipeline.named_steps['regressor']).__name__)])
                model_info.append(['Features Expected', str(self.cogs_pipeline.named_steps['preprocessor'].n_features_in_)])
            except:
                model_info.append(['Algorithm', 'Information not available'])

        if self.volume_pipeline:
            model_info.append(['', ''])
            model_info.append(['Volume Model Details', ''])
            try:
                model_info.append(['Algorithm', str(type(self.volume_pipeline.named_steps['regressor']).__name__)])
                model_info.append(['Features Expected', str(self.volume_pipeline.named_steps['preprocessor'].n_features_in_)])
            except:
                model_info.append(['Algorithm', 'Information not available'])

        model_info.append(['', ''])
        model_info.append(['Sales/COGS Feature List (Expected Order)', ''])

        # Add Sales/COGS feature list
        feature_order = [
            'Net_Sales_Volume_Litres', 'Marketing_Spend_EUR', 'Promotional_Event',
            'Consumer_Confidence_Index', 'Inflation_Rate_EUR', 'Avg_Temp_C',
            'Holiday_Indicator', 'Competitor_Activity_Index',
            'day_of_week', 'month', 'year', 'day_of_year', 'week_of_year',
            'Net_Sales_Revenue_EUR_lag1', 'COGS_EUR_lag1',
            'Country', 'Channel', 'Product_Category'
        ]

        for i, feature in enumerate(feature_order, 1):
            model_info.append([f'{i}.', feature])

        # Add volume model features
        model_info.append(['', ''])
        model_info.append(['Volume Model Features (Expected Order):', ''])

        volume_feature_order = [
            'Marketing_Spend_EUR', 'Promotional_Event',
            'Consumer_Confidence_Index', 'Inflation_Rate_EUR', 'Avg_Temp_C',
            'Holiday_Indicator', 'Competitor_Activity_Index',
            'day_of_week', 'month', 'year', 'day_of_year', 'week_of_year',
            'Net_Sales_Volume_Litres_lag1',
            'Country', 'Channel', 'Product_Category'
        ]

        for i, feature in enumerate(volume_feature_order, 1):
            model_info.append([f'{i}.', feature])

        df_model_info = pd.DataFrame(model_info, columns=['Category', 'Value'])
        df_model_info.to_excel(writer, sheet_name='Model Information', index=False)

    def _format_excel_file(self, filename):
        """Apply formatting to the Excel file"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = load_workbook(filename)

            # Define styles
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            title_font = Font(bold=True, size=14)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Format headers (first row)
                if ws.max_row > 0:
                    for cell in ws[1]:
                        if cell.value and str(cell.value).strip():
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center')

            wb.save(filename)
            print(f"✅ Excel formatting applied successfully")

        except Exception as e:
            print(f"⚠️ Could not apply Excel formatting: {e}")

    def _generate_recommendations(self):
        """Generate recommendations based on test results"""
        recommendations = []

        if 'holdout' in self.test_results:
            sales_mape = self.test_results['holdout']['sales']['MAPE']
            cogs_mape = self.test_results['holdout']['cogs']['MAPE']
            volume_mape = self.test_results['holdout']['volume']['MAPE']
            sales_r2 = self.test_results['holdout']['sales']['R2']
            cogs_r2 = self.test_results['holdout']['cogs']['R2']
            volume_r2 = self.test_results['holdout']['volume']['R2']

            if sales_mape > 15:
                recommendations.append("Sales model MAPE > 15% - Consider feature engineering or hyperparameter tuning")
            if cogs_mape > 15:
                recommendations.append("COGS model MAPE > 15% - May need additional training data or model complexity")
            if volume_mape > 15:
                recommendations.append("Volume model MAPE > 15% - Consider additional features, lagged variables, or model tuning")
            if sales_r2 < 0.7 or cogs_r2 < 0.7:
                recommendations.append("Low R² scores for Sales/COGS - Consider ensemble methods or advanced algorithms")
            if volume_r2 < 0.7:
                recommendations.append("Low R² score for Volume - Try more sophisticated models or feature engineering")

            if abs(self.test_results['holdout']['sales']['Bias_Percentage']) > 5:
                recommendations.append("Sales model shows significant bias - Review training data or add bias correction")
            if abs(self.test_results['holdout']['cogs']['Bias_Percentage']) > 5:
                recommendations.append("COGS model shows significant bias - Review training data or add bias correction")
            if abs(self.test_results['holdout']['volume']['Bias_Percentage']) > 5:
                recommendations.append("Volume model shows significant bias - Review training data or add bias correction")

        # General recommendations
        recommendations.extend([
            "Implement regular model retraining (monthly/quarterly)",
            "Monitor prediction accuracy on new data continuously", 
            "Consider A/B testing predictions against business forecasts",
            "Set up automated alerts for accuracy degradation",
            "Validate model performance across different market conditions"
        ])

        return recommendations[:10]  # Limit to top 10 recommendations


def main():
    """Main function to run all accuracy tests"""
    tester = AccuracyTester()

    try:
        # Load data and models
        tester.load_data_and_models()

        # Run all tests
        print("\n🚀 Starting comprehensive accuracy testing...")

        # 1. Holdout test accuracy
        tester.test_holdout_accuracy()

        # 2. Rolling window accuracy
        tester.test_rolling_window_accuracy()

        # 3. Segment-wise accuracy
        tester.test_segment_accuracy()

        # 4. Generate comprehensive report
        tester.generate_accuracy_report()

        # 5. Create visualizations
        tester.create_visualizations()

        # 6. Save results to Excel
        excel_filename = tester.save_results_to_excel()

        print(f"\n✅ Accuracy testing completed successfully!")
        print(f"📊 Detailed results saved to: {excel_filename}")

        return excel_filename

    except Exception as e:
        print(f"❌ Error during testing: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    main()